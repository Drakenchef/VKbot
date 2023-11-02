import requests
from bs4 import BeautifulSoup
import openpyxl
import re
import datetime

schedule_group_regex = re.compile(r'^[А-Я]{4}-\d{2}-\d{2}$')


def get_schedule():
    page = requests.get("https://www.mirea.ru/schedule/")
    soup = BeautifulSoup(page.text, "html.parser")
    result = soup.find(id='toggle-hl_2_1-hl_3_3').find_all('a', class_="uk-link-toggle")
    hrefs = []
    for i in result:
        href = i.get('href')
        if 'Zach' not in href and 'ekz' not in href:
            hrefs.append(href)
    schedule = dict()
    for href in hrefs:
        with open("file.xlsx", "wb") as f:
            resp = requests.get(href)  # запрос по ссылке
            f.write(resp.content)
        sheet = openpyxl.load_workbook('file.xlsx').active
        for cell in sheet['2']:
            if cell.value and schedule_group_regex.match(cell.value):
                column = cell.column
                group_schedule = [[], []]
                for i in range(4, 75, 14):
                    day_even = []
                    day_odd = []
                    for j in range(i, i + 12):
                        if j % 2 == 0:
                            if sheet.cell(row=j, column=column).value:
                                main_info = list()
                                for k in range(4):
                                    if sheet.cell(row=j, column=column + k).value:
                                        main_info.append(sheet.cell(row=j, column=column + k).value)
                            else:
                                main_info = []
                            day_odd.append(main_info)

                        else:
                            if sheet.cell(row=j, column=column).value:
                                main_info = list()
                                for k in range(4):
                                    if sheet.cell(row=j, column=column + k).value:
                                        main_info.append(sheet.cell(row=j, column=column + k).value)
                            else:
                                main_info = []
                            day_even.append(main_info)

                    group_schedule[0].append(day_odd)
                    group_schedule[1].append(day_even)
                schedule[cell.value.lower()] = group_schedule

    # for group in schedule:
    #     s = schedule[group]
    #     print(group)
    #     for day in range(len(s[0])):
    #         print('День ', day+1)
    #         for lesson in s[0][day]:
    #             print(lesson)

    return schedule


def day_schedule_to_string(lessons, lessons_date: datetime.date = None):
    months = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня', 'июля', 'августа', 'сентября', 'октября',
              'ноября', 'декабря']

    text = ''
    if lessons_date:
        text += f'Рассписание на {lessons_date.day} {months[lessons_date.month - 1]}:\n'

    for i in range(6):
        text += str(i + 1) + ') '
        main_info = lessons[i]
        if main_info and '\n' in main_info[0]:
            for j in range(len(main_info)):
                text += main_info[j].split("\n")[0] + ' '
            text += '\n   '
            for j in range(len(main_info)):
                text += main_info[j].split('\n')[2] + ' '
        elif main_info:
            for j in range(len(main_info)):
                text += main_info[j] + ' '
        else:
            text += '———'
        text += '\n'
    return text


if __name__ == '__main__':
    schedule = get_schedule()
    print(day_schedule_to_string(schedule['ИКБО-09-22'][1][2]))
